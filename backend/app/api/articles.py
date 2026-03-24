import uuid
import io
from datetime import datetime, date, timezone
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, update, func
from sqlalchemy.orm import selectinload
from pydantic import BaseModel

from app.core.database import get_db
from app.core.deps import get_current_user, require_admin_plus
from app.models.article import Article, ArticleStatus, ArticleModificationLog, Tonality
from app.models.article_flag import ArticleFlag
from app.models.article_read import ArticleRead
from app.models.client import Account, AccountRole, Client
from app.models.revue import Revue, Keyword
from app.services.nlp_service import nlp_service

router = APIRouter(prefix="/articles", tags=["articles"])


# ── Schemas ────────────────────────────────────────────────────────────────

class ArticleOut(BaseModel):
    id: uuid.UUID
    url: str
    source_domain: str | None
    title: str | None
    summary: str | None
    summary_en: str | None
    summary_ar: str | None
    author: str | None
    published_at: datetime | None
    tonality: Tonality | None
    tags: list | None
    relevance_score: float | None
    status: ArticleStatus
    scraping_error: str | None
    manually_entered: bool
    created_at: datetime
    # newspaper4k enriched fields
    image_url: str | None = None
    meta_description: str | None = None
    keyword_term: str | None = None   # résolu via la relation keyword
    keyword_query: str | None = None  # requête booléenne — pour le highlight côté UI
    content: str | None = None
    # Phase 5 — Intelligence Layer
    theme: str | None = None
    weak_signal: bool = False
    # Entités nommées (Phase 5)
    entities_persons: list | None = None
    entities_orgs: list | None = None
    entities_places: list | None = None
    # Intelligence approfondie (stockée depuis sprint 2)
    key_themes: list | None = None
    market_impact: str | None = None
    revue_id: uuid.UUID | None = None  # pour le picker de mots-clés en édition
    collection_method: str | None = None  # serpapi | rss | sitemap | manual
    reject_reason: str | None = None

    class Config:
        from_attributes = True

    @classmethod
    def from_article(cls, a: Article) -> "ArticleOut":
        """Sérialise un Article ORM en ArticleOut, en résolvant keyword_term et keyword_query."""
        d = {c: getattr(a, c) for c in cls.model_fields if hasattr(a, c)}
        d["keyword_term"] = a.keyword.term if (a.keyword is not None) else None
        d["keyword_query"] = a.keyword.query if (a.keyword is not None) else None
        return cls(**d)


class ArticleApprove(BaseModel):
    pass  # approuver tel quel


class ArticleModify(BaseModel):
    title: str | None = None
    summary: str | None = None
    tonality: Tonality | None = None
    tags: list[str] | None = None
    content: str | None = None
    # Phase 5 extensions
    author: str | None = None
    relevance_score: float | None = None   # 0.0 → 1.0
    weak_signal: bool | None = None
    theme: str | None = None
    keyword_id: uuid.UUID | None = None
    entities_persons: list[str] | None = None
    entities_orgs: list[str] | None = None


class ArticleReject(BaseModel):
    reason: str | None = None


class ManualEntry(BaseModel):
    title: str
    content: str
    author: str | None = None
    published_at: datetime | None = None
    summary: str | None = None


class AddByUrl(BaseModel):
    url: str
    keyword_id: uuid.UUID


class AssignFromFeed(BaseModel):
    rss_article_id: uuid.UUID
    keyword_id: uuid.UUID


class FixUrl(BaseModel):
    url: str


# ── Endpoints ──────────────────────────────────────────────────────────────

class BulkAction(BaseModel):
    ids: list[uuid.UUID]
    action: str  # "approve" | "reject" | "delete"
    reason: str | None = None


@router.post("/revue/{revue_id}/bulk")
async def bulk_article_action(
    revue_id: uuid.UUID,
    body: BulkAction,
    current_user: Account = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Actions en masse : approve, reject, delete sur une sélection d'articles."""
    if body.action not in ("approve", "reject", "delete"):
        raise HTTPException(400, "Action invalide — valeurs : approve | reject | delete")

    result = await db.execute(
        select(Article).where(
            Article.id.in_(body.ids),
            Article.revue_id == revue_id,
        )
    )
    articles = result.scalars().all()
    count = len(articles)

    if body.action == "approve":
        for a in articles:
            a.status = ArticleStatus.approved
            a.validated_by = current_user.id
            a.validated_at = datetime.utcnow()
        await db.commit()
    elif body.action == "reject":
        for a in articles:
            a.status = ArticleStatus.rejected
            a.validated_by = current_user.id
            a.validated_at = datetime.utcnow()
        await db.commit()
    elif body.action == "delete":
        for a in articles:
            await db.delete(a)
        await db.commit()

    return {"success": True, "count": count, "action": body.action}


@router.post("/revue/{revue_id}/bulk-analyze")
async def bulk_analyze_articles(
    revue_id: uuid.UUID,
    body: BulkAction,
    db: AsyncSession = Depends(get_db),
):
    """Relance l'analyse NLP (Claude) sur une sélection d'articles."""
    result = await db.execute(
        select(Article).where(
            Article.id.in_(body.ids),
            Article.revue_id == revue_id,
        ).options(selectinload(Article.keyword))
    )
    articles_to_analyze = result.scalars().all()
    analyzed = 0
    errors = 0

    for art in articles_to_analyze:
        if not art.content:
            errors += 1
            continue
        kw = art.keyword
        try:
            nlp = await nlp_service.analyze(
                title=art.title or "",
                content=art.content,
                url=art.url,
                keyword=kw.term if kw else "",
                scraper_author=art.author,
                scraper_date=art.published_at,
                meta_description=art.meta_description,
            )
            if not nlp.error:
                art.summary = nlp.summary
                art.summary_en = nlp.summary_en
                art.summary_ar = nlp.summary_ar
                art.tonality = nlp.tonality
                art.tags = nlp.tags
                art.relevance_score = nlp.relevance_score
                art.theme = nlp.theme
                art.weak_signal = nlp.weak_signal
                art.entities_persons = nlp.entities_persons or []
                art.entities_orgs = nlp.entities_orgs or []
                art.entities_places = nlp.entities_places or []
                art.key_themes = nlp.key_themes or []
                art.market_impact = nlp.market_impact
                analyzed += 1
            else:
                errors += 1
        except Exception:
            errors += 1

    await db.commit()
    return {"analyzed": analyzed, "errors": errors, "total": len(articles_to_analyze)}


@router.get("/revue/{revue_id}/analytics")
async def revue_analytics(
    revue_id: uuid.UUID,
    days: int = Query(30, ge=7, le=90),
    db: AsyncSession = Depends(get_db),
):
    """Analytics pour le dashboard: volume collecté, tonalité, top sources."""
    from datetime import timedelta
    from sqlalchemy import cast, Date as SADate

    since = datetime.now(timezone.utc) - timedelta(days=days)

    # Volume par jour (tous articles collectés)
    volume_rows = await db.execute(
        select(
            cast(Article.created_at, SADate).label("day"),
            func.count().label("count"),
        )
        .where(Article.revue_id == revue_id, Article.created_at >= since)
        .group_by("day")
        .order_by("day")
    )
    volume = [{"date": str(row.day), "count": row.count} for row in volume_rows]

    # Distribution tonalité (articles validés)
    tonality_rows = await db.execute(
        select(Article.tonality, func.count().label("count"))
        .where(
            Article.revue_id == revue_id,
            Article.status.in_([ArticleStatus.approved, ArticleStatus.modified]),
            Article.tonality.isnot(None),
        )
        .group_by(Article.tonality)
    )
    tonality = [
        {"tonality": row.tonality.value if row.tonality else "unknown", "count": row.count}
        for row in tonality_rows
    ]

    validated_filter = [
        Article.revue_id == revue_id,
        Article.status.in_([ArticleStatus.approved, ArticleStatus.modified]),
    ]

    # Top 8 sources (articles validés)
    sources_rows = await db.execute(
        select(Article.source_domain, func.count().label("count"))
        .where(*validated_filter, Article.source_domain.isnot(None))
        .group_by(Article.source_domain)
        .order_by(func.count().desc())
        .limit(8)
    )
    sources = [{"domain": row.source_domain, "count": row.count} for row in sources_rows]

    # Volume validé par jour
    volume_validated_rows = await db.execute(
        select(
            cast(Article.created_at, SADate).label("day"),
            func.count().label("count"),
        )
        .where(*validated_filter, Article.created_at >= since)
        .group_by("day").order_by("day")
    )
    volume_validated = [{"date": str(r.day), "count": r.count} for r in volume_validated_rows]

    # Top keywords (articles validés)
    keyword_rows = await db.execute(
        select(Keyword.term, func.count().label("count"))
        .join(Article, Article.keyword_id == Keyword.id)
        .where(*validated_filter)
        .group_by(Keyword.term)
        .order_by(func.count().desc())
        .limit(10)
    )
    keywords = [{"term": row.term, "count": row.count} for row in keyword_rows]

    # Top auteurs / journalistes (articles validés)
    author_rows = await db.execute(
        select(Article.author, func.count().label("count"))
        .where(*validated_filter, Article.author.isnot(None))
        .group_by(Article.author)
        .order_by(func.count().desc())
        .limit(8)
    )
    authors = [{"author": row.author, "count": row.count} for row in author_rows]

    # Langues (depuis keyword.language, articles validés)
    language_rows = await db.execute(
        select(Keyword.language, func.count().label("count"))
        .join(Article, Article.keyword_id == Keyword.id)
        .where(*validated_filter)
        .group_by(Keyword.language)
        .order_by(func.count().desc())
    )
    languages = [{"language": row.language, "count": row.count} for row in language_rows]

    return {
        "volume": volume,
        "volume_validated": volume_validated,
        "tonality": tonality,
        "sources": sources,
        "keywords": keywords,
        "authors": authors,
        "languages": languages,
    }


# ── Suggestions autocomplete pour l'édition ──────────────────────────────

@router.get("/revue/{revue_id}/suggestions")
async def article_suggestions(
    revue_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
):
    """
    Retourne les tags, personnalités et institutions les plus utilisés
    dans les articles validés d'une revue — pour l'autocomplete en édition.
    """
    result = await db.execute(
        select(Article.tags, Article.entities_persons, Article.entities_orgs)
        .where(
            Article.revue_id == revue_id,
            Article.status.in_([ArticleStatus.approved, ArticleStatus.modified]),
        )
        .limit(500)
    )
    rows = result.all()

    tags_cnt: dict[str, int] = {}
    persons_cnt: dict[str, int] = {}
    orgs_cnt: dict[str, int] = {}

    for row in rows:
        for t in (row.tags or []):
            tags_cnt[t] = tags_cnt.get(t, 0) + 1
        for p in (row.entities_persons or []):
            persons_cnt[p] = persons_cnt.get(p, 0) + 1
        for o in (row.entities_orgs or []):
            orgs_cnt[o] = orgs_cnt.get(o, 0) + 1

    def top(d: dict, n: int = 100) -> list[str]:
        return [k for k, _ in sorted(d.items(), key=lambda x: -x[1])[:n]]

    return {
        "tags": top(tags_cnt),
        "persons": top(persons_cnt),
        "orgs": top(orgs_cnt),
    }


# ── Rapport statistiques mensuel ─────────────────────────────────────────

@router.get("/revue/{revue_id}/monthly-report")
async def revue_monthly_report(
    revue_id: uuid.UUID,
    date_from: str = Query(..., description="Date de début ISO (YYYY-MM-DD)"),
    date_to: str = Query(..., description="Date de fin ISO (YYYY-MM-DD)"),
    db: AsyncSession = Depends(get_db),
):
    """Rapport statistiques complet pour une période : KPIs, tonalité, sources, auteurs."""
    from datetime import timedelta
    from sqlalchemy import cast, Date as SADate, extract

    # ─── Parse dates ──────────────────────────────────────────────────────
    try:
        d_from_obj = date.fromisoformat(date_from)
        d_to_obj   = date.fromisoformat(date_to)
    except ValueError:
        raise HTTPException(400, "Format de date invalide — attendu YYYY-MM-DD")

    d_from     = datetime(d_from_obj.year, d_from_obj.month, d_from_obj.day, tzinfo=timezone.utc)
    d_to       = datetime(d_to_obj.year, d_to_obj.month, d_to_obj.day, 23, 59, 59, tzinfo=timezone.utc)
    days_count = (d_to_obj - d_from_obj).days + 1

    # ─── Revue + client ───────────────────────────────────────────────────
    revue_row = (await db.execute(
        select(Revue, Client)
        .join(Client, Revue.client_id == Client.id)
        .where(Revue.id == revue_id)
    )).first()
    if not revue_row:
        raise HTTPException(404, "Revue non trouvée")
    revue_obj, client_obj = revue_row

    # ─── Shared expressions ───────────────────────────────────────────────
    date_expr = func.coalesce(Article.published_at, Article.created_at)
    day_col   = cast(date_expr, SADate)

    base_filter = [
        Article.revue_id == revue_id,
        Article.status.in_([ArticleStatus.approved, ArticleStatus.modified]),
        date_expr >= d_from,
        date_expr <= d_to,
    ]

    # ─── Totals ───────────────────────────────────────────────────────────
    total = (await db.execute(select(func.count()).where(*base_filter))).scalar() or 0

    # Période précédente (même durée, juste avant)
    delta      = timedelta(days=days_count)
    prev_from  = d_from - delta
    prev_to    = d_from - timedelta(seconds=1)
    prev_total = (await db.execute(
        select(func.count()).where(
            Article.revue_id == revue_id,
            Article.status.in_([ArticleStatus.approved, ArticleStatus.modified]),
            date_expr >= prev_from,
            date_expr <= prev_to,
        )
    )).scalar() or 0

    if prev_total == 0:
        evolution_pct = 100.0 if total > 0 else 0.0
    else:
        evolution_pct = round((total - prev_total) / prev_total * 100, 1)

    avg_per_day = round(total / days_count, 1) if days_count > 0 else 0.0

    # ─── Volume journalier ────────────────────────────────────────────────
    daily_map = {
        str(r.day): r.count
        for r in (await db.execute(
            select(day_col.label("day"), func.count().label("count"))
            .where(*base_filter)
            .group_by(day_col)
            .order_by(day_col)
        )).all()
    }
    daily_volume = [
        {
            "date": str(d_from_obj + timedelta(days=i)),
            "count": daily_map.get(str(d_from_obj + timedelta(days=i)), 0),
        }
        for i in range(days_count)
    ]

    # Journée pic
    peak_entry  = max(daily_volume, key=lambda x: x["count"]) if daily_volume else {"date": date_from, "count": 0}
    WEEKDAYS_FR = ["lundi", "mardi", "mercredi", "jeudi", "vendredi", "samedi", "dimanche"]
    peak_weekday = WEEKDAYS_FR[date.fromisoformat(peak_entry["date"]).weekday()]

    # ─── Historique 12 mois ───────────────────────────────────────────────
    hist_from = d_to - timedelta(days=366)
    hist_map = {
        (int(r.yr), int(r.mo)): r.count
        for r in (await db.execute(
            select(
                extract("year",  date_expr).label("yr"),
                extract("month", date_expr).label("mo"),
                func.count().label("count"),
            )
            .where(
                Article.revue_id == revue_id,
                Article.status.in_([ArticleStatus.approved, ArticleStatus.modified]),
                date_expr >= hist_from,
                date_expr <= d_to,
            )
            .group_by("yr", "mo")
            .order_by("yr", "mo")
        )).all()
    }

    monthly_history = []
    for i in range(11, -1, -1):
        mo = d_to_obj.month - i
        yr = d_to_obj.year
        while mo <= 0:
            mo += 12
            yr -= 1
        monthly_history.append({"month": f"{yr:04d}-{mo:02d}", "count": hist_map.get((yr, mo), 0)})

    avg_per_month = round(sum(m["count"] for m in monthly_history) / 12, 1)
    peak_month    = max(monthly_history, key=lambda x: x["count"])

    # ─── Langues ──────────────────────────────────────────────────────────
    languages = {
        r.language: r.count
        for r in (await db.execute(
            select(Keyword.language, func.count().label("count"))
            .join(Article, Article.keyword_id == Keyword.id)
            .where(*base_filter)
            .group_by(Keyword.language)
        )).all()
    }

    # ─── Tonalité ─────────────────────────────────────────────────────────
    tone_map = {
        r.tonality.value: r.count
        for r in (await db.execute(
            select(Article.tonality, func.count().label("count"))
            .where(*base_filter, Article.tonality.isnot(None))
            .group_by(Article.tonality)
        )).all()
    }

    def tone_stat(key: str) -> dict:
        cnt = tone_map.get(key, 0)
        return {
            "count": cnt,
            "pct": round(cnt / total * 100, 1) if total > 0 else 0.0,
            "avg_per_day": round(cnt / days_count, 1) if days_count > 0 else 0.0,
        }

    tonality_out = {
        "positive": tone_stat("positive"),
        "negative": tone_stat("negative"),
        "neutral":  tone_stat("neutral"),
    }

    # ─── Tonalité journalière ─────────────────────────────────────────────
    daily_tone_raw = (await db.execute(
        select(day_col.label("day"), Article.tonality, func.count().label("count"))
        .where(*base_filter, Article.tonality.isnot(None))
        .group_by(day_col, Article.tonality)
        .order_by(day_col)
    )).all()

    daily_tone_map: dict[str, dict] = {}
    for r in daily_tone_raw:
        k = str(r.day)
        if k not in daily_tone_map:
            daily_tone_map[k] = {"positive": 0, "negative": 0, "neutral": 0}
        daily_tone_map[k][r.tonality.value] = r.count

    daily_tonality = [
        {
            "date": dv["date"],
            "positive": daily_tone_map.get(dv["date"], {}).get("positive", 0),
            "negative": daily_tone_map.get(dv["date"], {}).get("negative", 0),
            "neutral":  daily_tone_map.get(dv["date"], {}).get("neutral", 0),
        }
        for dv in daily_volume
    ]

    # ─── Tonalité historique mensuel ──────────────────────────────────────
    monthly_tone_raw = (await db.execute(
        select(
            extract("year",  date_expr).label("yr"),
            extract("month", date_expr).label("mo"),
            Article.tonality,
            func.count().label("count"),
        )
        .where(
            Article.revue_id == revue_id,
            Article.status.in_([ArticleStatus.approved, ArticleStatus.modified]),
            Article.tonality.isnot(None),
            date_expr >= hist_from,
            date_expr <= d_to,
        )
        .group_by("yr", "mo", Article.tonality)
    )).all()

    monthly_tone_map: dict[tuple, dict] = {}
    for r in monthly_tone_raw:
        key = (int(r.yr), int(r.mo))
        if key not in monthly_tone_map:
            monthly_tone_map[key] = {"positive": 0, "negative": 0, "neutral": 0}
        monthly_tone_map[key][r.tonality.value] = r.count

    monthly_tonality_history = []
    for m in monthly_history:
        yr_s, mo_s = m["month"].split("-")
        tm = monthly_tone_map.get((int(yr_s), int(mo_s)), {"positive": 0, "negative": 0, "neutral": 0})
        monthly_tonality_history.append({
            "month": m["month"],
            "positive": tm.get("positive", 0),
            "negative": tm.get("negative", 0),
            "neutral":  tm.get("neutral", 0),
        })

    # ─── Top sources ──────────────────────────────────────────────────────
    async def _top_sources(tone: Tonality) -> list:
        rows = (await db.execute(
            select(Article.source_domain, func.count().label("count"))
            .where(*base_filter, Article.tonality == tone, Article.source_domain.isnot(None))
            .group_by(Article.source_domain)
            .order_by(func.count().desc())
            .limit(10)
        )).all()
        return [
            {
                "domain": r.source_domain,
                "count": r.count,
                "pct": round(r.count / total * 100, 1) if total > 0 else 0.0,
            }
            for r in rows
        ]

    # ─── Top auteurs ──────────────────────────────────────────────────────
    _UNKNOWN_AUTHORS = {"unknown", "inconnu", "n/a", "none", "null", "-", "", "<unknown>"}

    def _is_unknown_author(name: str) -> bool:
        n = (name or "").strip().lower()
        return n in _UNKNOWN_AUTHORS or "unknown" in n or not n

    async def _top_authors(tone: Tonality) -> list:
        rows = (await db.execute(
            select(Article.author, Article.source_domain, func.count().label("count"))
            .where(*base_filter, Article.tonality == tone, Article.author.isnot(None))
            .group_by(Article.author, Article.source_domain)
            .order_by(func.count().desc())
            .limit(20)  # fetch more to account for filtered unknowns
        )).all()
        results = []
        for r in rows:
            if _is_unknown_author(r.author or ""):
                continue
            results.append({
                "author": r.author,
                "source": r.source_domain or "",
                "count": r.count,
                "pct": round(r.count / total * 100, 1) if total > 0 else 0.0,
            })
            if len(results) >= 10:
                break
        return results

    top_sources_positive = await _top_sources(Tonality.positive)
    top_sources_negative = await _top_sources(Tonality.negative)
    top_authors_positive = await _top_authors(Tonality.positive)
    top_authors_negative = await _top_authors(Tonality.negative)

    # ─── Label période ────────────────────────────────────────────────────
    MONTHS_FR = [
        "janvier", "février", "mars", "avril", "mai", "juin",
        "juillet", "août", "septembre", "octobre", "novembre", "décembre",
    ]
    # Si c'est un mois calendaire complet → "Mars 2026", sinon → plage
    import calendar as _cal
    last_day = _cal.monthrange(d_from_obj.year, d_from_obj.month)[1]
    if (
        d_from_obj.day == 1
        and d_to_obj.year == d_from_obj.year
        and d_to_obj.month == d_from_obj.month
        and d_to_obj.day == last_day
    ):
        label = f"{MONTHS_FR[d_from_obj.month - 1].capitalize()} {d_from_obj.year}"
    else:
        label = (
            f"{d_from_obj.day:02d}/{d_from_obj.month:02d}/{d_from_obj.year}"
            f" → "
            f"{d_to_obj.day:02d}/{d_to_obj.month:02d}/{d_to_obj.year}"
        )

    return {
        "period": {
            "date_from": date_from,
            "date_to": date_to,
            "label": label,
            "days": days_count,
        },
        "revue_name": revue_obj.name,
        "client_name": client_obj.name,
        "total": total,
        "avg_per_day": avg_per_day,
        "evolution_pct": evolution_pct,
        "peak_day": {
            "date": peak_entry["date"],
            "weekday": peak_weekday,
            "count": peak_entry["count"],
        },
        "daily_volume": daily_volume,
        "monthly_history": monthly_history,
        "avg_per_month": avg_per_month,
        "peak_month": peak_month,
        "languages": languages,
        "tonality": tonality_out,
        "daily_tonality": daily_tonality,
        "monthly_tonality_history": monthly_tonality_history,
        "top_sources_positive": top_sources_positive,
        "top_sources_negative": top_sources_negative,
        "top_authors_positive": top_authors_positive,
        "top_authors_negative": top_authors_negative,
    }


# ── Read / Star tracking ────────────────────────────────────────────────

@router.get("/revue/{revue_id}/reads")
async def get_read_map(
    revue_id: uuid.UUID,
    current_user: Account = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Retourne la map lu/étoile pour tous les articles d'une revue, pour l'utilisateur courant."""
    rows = await db.execute(
        select(ArticleRead)
        .where(
            ArticleRead.user_id == current_user.id,
            ArticleRead.article_id.in_(
                select(Article.id).where(Article.revue_id == revue_id)
            ),
        )
    )
    result = {}
    for r in rows.scalars().all():
        result[str(r.article_id)] = {
            "is_read": r.read_at is not None,
            "is_starred": r.starred,
            "read_at": r.read_at.isoformat() if r.read_at else None,
        }
    return result


@router.get("/revue/{revue_id}/unread-count")
async def get_unread_count(
    revue_id: uuid.UUID,
    current_user: Account = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Nombre d'articles validés non lus et d'articles étoilés pour l'utilisateur courant."""
    # Total articles validés de la revue
    total = await db.scalar(
        select(func.count()).where(
            Article.revue_id == revue_id,
            Article.status.in_([ArticleStatus.approved, ArticleStatus.modified]),
        )
    ) or 0

    # Articles déjà lus
    read_count = await db.scalar(
        select(func.count()).where(
            ArticleRead.user_id == current_user.id,
            ArticleRead.read_at.isnot(None),
            ArticleRead.article_id.in_(
                select(Article.id).where(
                    Article.revue_id == revue_id,
                    Article.status.in_([ArticleStatus.approved, ArticleStatus.modified]),
                )
            ),
        )
    ) or 0

    # Articles étoilés
    starred_count = await db.scalar(
        select(func.count()).where(
            ArticleRead.user_id == current_user.id,
            ArticleRead.starred == True,
            ArticleRead.article_id.in_(
                select(Article.id).where(
                    Article.revue_id == revue_id,
                    Article.status.in_([ArticleStatus.approved, ArticleStatus.modified]),
                )
            ),
        )
    ) or 0

    return {"total": total, "unread": total - read_count, "starred": starred_count}


@router.post("/{article_id}/read")
async def mark_read(
    article_id: uuid.UUID,
    current_user: Account = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Marque un article comme lu (upsert)."""
    existing = await db.scalar(
        select(ArticleRead).where(
            ArticleRead.user_id == current_user.id,
            ArticleRead.article_id == article_id,
        )
    )
    if existing:
        if existing.read_at is None:
            existing.read_at = datetime.now(timezone.utc)
            await db.commit()
    else:
        db.add(ArticleRead(
            user_id=current_user.id,
            article_id=article_id,
            read_at=datetime.now(timezone.utc),
        ))
        await db.commit()
    return {"status": "ok"}


@router.delete("/{article_id}/read")
async def mark_unread(
    article_id: uuid.UUID,
    current_user: Account = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Marque un article comme non lu."""
    existing = await db.scalar(
        select(ArticleRead).where(
            ArticleRead.user_id == current_user.id,
            ArticleRead.article_id == article_id,
        )
    )
    if existing:
        existing.read_at = None
        await db.commit()
    return {"status": "ok"}


@router.post("/{article_id}/star")
async def toggle_star(
    article_id: uuid.UUID,
    current_user: Account = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Bascule l'état étoilé d'un article."""
    existing = await db.scalar(
        select(ArticleRead).where(
            ArticleRead.user_id == current_user.id,
            ArticleRead.article_id == article_id,
        )
    )
    if existing:
        existing.starred = not existing.starred
        existing.starred_at = datetime.now(timezone.utc) if existing.starred else None
    else:
        existing = ArticleRead(
            user_id=current_user.id,
            article_id=article_id,
            starred=True,
            starred_at=datetime.now(timezone.utc),
        )
        db.add(existing)
    await db.commit()
    return {"starred": existing.starred}


@router.post("/revue/{revue_id}/read-all")
async def mark_all_read(
    revue_id: uuid.UUID,
    current_user: Account = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Marque tous les articles validés d'une revue comme lus."""
    # Récupère les IDs non encore lus
    all_ids_rows = await db.execute(
        select(Article.id).where(
            Article.revue_id == revue_id,
            Article.status.in_([ArticleStatus.approved, ArticleStatus.modified]),
        )
    )
    all_ids = {row[0] for row in all_ids_rows}

    already_rows = await db.execute(
        select(ArticleRead).where(
            ArticleRead.user_id == current_user.id,
            ArticleRead.article_id.in_(all_ids),
        )
    )
    already = {r.article_id: r for r in already_rows.scalars().all()}

    now = datetime.now(timezone.utc)
    count = 0
    for aid in all_ids:
        if aid in already:
            if already[aid].read_at is None:
                already[aid].read_at = now
                count += 1
        else:
            db.add(ArticleRead(user_id=current_user.id, article_id=aid, read_at=now))
            count += 1

    await db.commit()
    return {"count": count}

@router.get("/revue/{revue_id}/stats")
async def revue_stats(revue_id: uuid.UUID, db: AsyncSession = Depends(get_db)):
    """Compteurs pour le dashboard HITL."""
    today_start = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0)

    rows = await db.execute(
        select(Article.status, func.count()).where(Article.revue_id == revue_id).group_by(Article.status)
    )
    counts = {status: n for status, n in rows}

    validated_today = await db.scalar(
        select(func.count()).where(
            Article.revue_id == revue_id,
            Article.status.in_([ArticleStatus.approved, ArticleStatus.modified]),
            Article.validated_at >= today_start,
        )
    )

    return {
        "pending": counts.get(ArticleStatus.pending, 0),
        "error": counts.get(ArticleStatus.error, 0),
        "validated_today": validated_today or 0,
        "approved": counts.get(ArticleStatus.approved, 0),
        "modified": counts.get(ArticleStatus.modified, 0),
        "rejected": counts.get(ArticleStatus.rejected, 0),
    }


@router.get("/revue/{revue_id}", response_model=list[ArticleOut])
async def list_articles(
    revue_id: uuid.UUID,
    status: ArticleStatus | None = Query(None),
    tonality: Tonality | None = Query(None),
    source_domain: str | None = Query(None),
    date_from: date | None = Query(None),
    date_to: date | None = Query(None),
    min_relevance: float | None = Query(None, ge=0, le=100),
    keyword_id: uuid.UUID | None = Query(None),
    theme: str | None = Query(None),
    weak_signal: bool | None = Query(None),
    current_user: Account = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Liste les articles d'une revue avec filtres optionnels."""
    stmt = (
        select(Article)
        .where(Article.revue_id == revue_id)
        .options(selectinload(Article.keyword))
    )
    if status:
        stmt = stmt.where(Article.status == status)
    if tonality:
        stmt = stmt.where(Article.tonality == tonality)
    if source_domain:
        stmt = stmt.where(Article.source_domain.ilike(f"%{source_domain}%"))
    if date_from:
        stmt = stmt.where(Article.published_at >= datetime.combine(date_from, datetime.min.time()))
    if date_to:
        stmt = stmt.where(Article.published_at <= datetime.combine(date_to, datetime.max.time()))
    if min_relevance is not None:
        stmt = stmt.where(Article.relevance_score >= min_relevance / 100)
    if keyword_id is not None:
        stmt = stmt.where(Article.keyword_id == keyword_id)
    if theme:
        stmt = stmt.where(Article.theme == theme)
    if weak_signal is not None:
        stmt = stmt.where(Article.weak_signal == weak_signal)
    stmt = stmt.order_by(Article.created_at.desc())

    result = await db.execute(stmt)
    articles = [ArticleOut.from_article(a) for a in result.scalars().all()]

    # Masquer le texte intégral pour les utilisateurs côté client
    is_client = current_user.role in (AccountRole.client_admin, AccountRole.client_user)
    if is_client:
        for a in articles:
            a.content = None

    return articles


@router.get("/revue/{revue_id}/export")
async def export_articles(
    revue_id: uuid.UUID,
    fmt: str = Query("excel", pattern="^(excel|pdf)$"),
    status: ArticleStatus | None = Query(None),
    current_user: Account = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Export des articles validés en Excel ou PDF."""
    stmt = select(Article).where(Article.revue_id == revue_id)
    if status:
        stmt = stmt.where(Article.status == status)
    else:
        stmt = stmt.where(Article.status.in_([ArticleStatus.approved, ArticleStatus.modified]))
    stmt = stmt.order_by(Article.published_at.desc().nulls_last())

    result = await db.execute(stmt)
    articles = result.scalars().all()

    revue = await db.get(Revue, revue_id)
    revue_name = revue.name if revue else "Revue"
    export_date = datetime.now().strftime("%Y-%m-%d")

    if fmt == "excel":
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = Workbook()
        ws = wb.active
        ws.title = "Articles validés"

        # Header
        headers = ["Titre", "Source", "Auteur", "Date publication", "Tonalité", "Pertinence", "Tags", "Résumé", "URL", "Statut"]
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="1E3A5F")

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Rows
        for row, a in enumerate(articles, 2):
            ws.cell(row=row, column=1, value=a.title or "—")
            ws.cell(row=row, column=2, value=a.source_domain or "—")
            ws.cell(row=row, column=3, value=a.author or "—")
            ws.cell(row=row, column=4, value=a.published_at.strftime("%d/%m/%Y") if a.published_at else "—")
            ws.cell(row=row, column=5, value=a.tonality.value if a.tonality else "—")
            ws.cell(row=row, column=6, value=f"{a.relevance_score * 100:.0f}%" if a.relevance_score else "—")
            ws.cell(row=row, column=7, value=", ".join(a.tags) if a.tags else "—")
            ws.cell(row=row, column=8, value=a.summary or "—")
            ws.cell(row=row, column=9, value=a.url)
            ws.cell(row=row, column=10, value=a.status.value)

        # Column widths
        for col, width in zip(range(1, 11), [40, 20, 20, 15, 12, 12, 30, 60, 50, 12]):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        filename = f"revue_{revue_name}_{export_date}.xlsx"
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    else:  # PDF
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=2*cm, rightMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)
        styles = getSampleStyleSheet()
        story = []

        title_style = ParagraphStyle("title", parent=styles["Title"], fontSize=18, textColor=colors.HexColor("#1E3A5F"))
        subtitle_style = ParagraphStyle("sub", parent=styles["Normal"], fontSize=10, textColor=colors.grey)
        item_title_style = ParagraphStyle("ititle", parent=styles["Heading2"], fontSize=11, spaceAfter=4)
        body_style = ParagraphStyle("body", parent=styles["Normal"], fontSize=9, leading=13)
        meta_style = ParagraphStyle("meta", parent=styles["Normal"], fontSize=8, textColor=colors.HexColor("#666666"))

        story.append(Paragraph(f"Revue de Presse — {revue_name}", title_style))
        story.append(Paragraph(f"Exporté le {export_date} · {len(articles)} article(s) validé(s)", subtitle_style))
        story.append(Spacer(1, 0.5*cm))

        TONE_COLOR = {"positive": "#16a34a", "negative": "#dc2626", "neutral": "#6b7280", "mixed": "#d97706"}

        for a in articles:
            tone_color = TONE_COLOR.get(a.tonality.value if a.tonality else "neutral", "#6b7280")
            story.append(Paragraph(a.title or "Sans titre", item_title_style))

            meta_parts = []
            if a.source_domain: meta_parts.append(a.source_domain)
            if a.published_at: meta_parts.append(a.published_at.strftime("%d/%m/%Y"))
            if a.author: meta_parts.append(a.author)
            if a.tonality: meta_parts.append(f'<font color="{tone_color}">■ {a.tonality.value}</font>')
            if a.relevance_score: meta_parts.append(f"Pertinence: {a.relevance_score * 100:.0f}%")
            story.append(Paragraph(" · ".join(meta_parts), meta_style))

            if a.summary:
                story.append(Spacer(1, 0.2*cm))
                story.append(Paragraph(a.summary, body_style))

            if a.tags:
                story.append(Paragraph(f"Tags: {', '.join(a.tags)}", meta_style))

            story.append(Spacer(1, 0.5*cm))

        doc.build(story)
        buf.seek(0)
        filename = f"revue_{revue_name}_{export_date}.pdf"
        return StreamingResponse(
            buf,
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )


@router.post("/revue/{revue_id}/add-by-url", response_model=ArticleOut)
async def add_article_by_url(
    revue_id: uuid.UUID,
    body: AddByUrl,
    current_user: Account = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Crée un article à partir d'une URL :
    1. Scrape avec newspaper4k (titre, contenu, auteur, date, image, meta)
    2. Analyse NLP (résumé FR/EN/AR, tonalité, tags, pertinence)
    3. Sauvegarde et retourne l'article complet prêt à valider.
    """
    from urllib.parse import urlparse
    from app.services.scraper_service import scraper_service
    from app.models.revue import RevueKeyword

    # Vérifier que le keyword est bien associé à cette revue (via RevueKeyword)
    rk_result = await db.execute(
        select(RevueKeyword).where(
            RevueKeyword.revue_id == revue_id,
            RevueKeyword.keyword_id == body.keyword_id,
        ).limit(1)
    )
    rk = rk_result.scalar_one_or_none()
    if not rk:
        raise HTTPException(status_code=404, detail="Keyword introuvable pour cette revue")

    # Récupérer le keyword
    kw = await db.get(Keyword, body.keyword_id)
    if not kw:
        raise HTTPException(status_code=404, detail="Keyword introuvable")

    # Extraire le domain
    source_domain = urlparse(body.url).netloc or None

    # 1. Scraping newspaper4k (avec la langue du keyword)
    scraped = await scraper_service.scrape(body.url, language=kw.language or "fr")

    # Créer l'article (même si scraping échoué — on garde l'URL)
    article = Article(
        revue_id=revue_id,
        keyword_id=body.keyword_id,
        url=body.url,
        source_domain=source_domain,
        manually_entered=True,
    )

    if not scraped.success:
        article.status = ArticleStatus.error
        article.scraping_error = scraped.error
        db.add(article)
        await db.commit()
        await db.refresh(article)
        # Recharger avec keyword
        result = await db.execute(
            select(Article).where(Article.id == article.id).options(selectinload(Article.keyword))
        )
        return ArticleOut.from_article(result.scalar_one())

    # Remplir les champs scraper
    article.title = scraped.title
    article.content = scraped.content
    article.author = scraped.author
    article.published_at = scraped.published_at
    article.image_url = scraped.image_url
    article.meta_description = scraped.meta_description

    # 2. Analyse NLP
    nlp = await nlp_service.analyze(
        title=scraped.title or "",
        content=scraped.content or "",
        url=body.url,
        keyword=kw.term,
        scraper_author=scraped.author,
        scraper_date=scraped.published_at,
        meta_description=scraped.meta_description,
    )

    if not nlp.error:
        article.summary = nlp.summary
        article.summary_en = nlp.summary_en
        article.summary_ar = nlp.summary_ar
        article.tonality = nlp.tonality
        article.tags = nlp.tags
        article.relevance_score = nlp.relevance_score
        article.theme = nlp.theme
        article.weak_signal = nlp.weak_signal
        article.entities_persons = nlp.entities_persons or []
        article.entities_orgs = nlp.entities_orgs or []
        article.entities_places = nlp.entities_places or []
        article.key_themes = nlp.key_themes or []
        article.market_impact = nlp.market_impact
        if nlp.author and not article.author:
            article.author = nlp.author
        if nlp.published_at and not article.published_at:
            article.published_at = nlp.published_at

    article.status = ArticleStatus.pending
    db.add(article)
    await db.commit()

    # Recharger avec keyword pour la sérialisation
    result = await db.execute(
        select(Article).where(Article.id == article.id).options(selectinload(Article.keyword))
    )
    return ArticleOut.from_article(result.scalar_one())


@router.post("/revue/{revue_id}/assign-from-feed", response_model=ArticleOut)
async def assign_article_from_feed(
    revue_id: uuid.UUID,
    body: AssignFromFeed,
    current_user: Account = Depends(require_admin_plus),
    db: AsyncSession = Depends(get_db),
):
    """
    Affecte manuellement un article du fil RSS à une revue.
    Réutilise les données déjà extraites (pas de re-scraping), puis lance l'analyse NLP.
    """
    from urllib.parse import urlparse
    from app.models.revue import RevueKeyword
    from app.models.rss_article import RssArticle

    # Vérifier que le keyword appartient bien à cette revue
    rk_result = await db.execute(
        select(RevueKeyword).where(
            RevueKeyword.revue_id == revue_id,
            RevueKeyword.keyword_id == body.keyword_id,
        ).limit(1)
    )
    if not rk_result.scalar_one_or_none():
        raise HTTPException(404, "Keyword introuvable pour cette revue")

    kw = await db.get(Keyword, body.keyword_id)
    if not kw:
        raise HTTPException(404, "Keyword introuvable")

    # Récupérer l'article RSS source
    rss = await db.get(RssArticle, body.rss_article_id)
    if not rss:
        raise HTTPException(404, "Article introuvable dans le fil")

    # Détecter un doublon URL sur la même revue
    dup = await db.execute(
        select(Article).where(
            Article.revue_id == revue_id,
            Article.url == rss.url,
        ).limit(1)
    )
    if dup.scalar_one_or_none():
        raise HTTPException(409, "Cet article est déjà présent dans cette revue")

    source_domain = urlparse(rss.url).netloc or None

    article = Article(
        revue_id=revue_id,
        keyword_id=body.keyword_id,
        url=rss.url,
        source_domain=source_domain,
        title=rss.title,
        content=rss.content,
        author=rss.author,
        published_at=rss.published_at,
        image_url=rss.image_url,
        manually_entered=True,
        collection_method="manual",
    )

    # Analyse NLP sur le contenu déjà extrait
    if rss.title or rss.content:
        nlp = await nlp_service.analyze(
            title=rss.title or "",
            content=rss.content or "",
            url=rss.url,
            keyword=kw.term,
            scraper_author=rss.author,
            scraper_date=rss.published_at,
            meta_description=rss.summary,
        )
        if not nlp.error:
            article.summary = nlp.summary
            article.summary_en = nlp.summary_en
            article.summary_ar = nlp.summary_ar
            article.tonality = nlp.tonality
            article.tags = nlp.tags
            article.relevance_score = nlp.relevance_score
            article.theme = nlp.theme
            article.weak_signal = nlp.weak_signal
            article.entities_persons = nlp.entities_persons or []
            article.entities_orgs = nlp.entities_orgs or []
            article.entities_places = nlp.entities_places or []
            article.key_themes = nlp.key_themes or []
            article.market_impact = nlp.market_impact
            if nlp.author and not article.author:
                article.author = nlp.author

    article.status = ArticleStatus.pending
    db.add(article)
    await db.commit()

    result = await db.execute(
        select(Article).where(Article.id == article.id).options(selectinload(Article.keyword))
    )
    return ArticleOut.from_article(result.scalar_one())


@router.patch("/{article_id}/fix-url", response_model=ArticleOut)
async def fix_article_url(
    article_id: uuid.UUID,
    body: FixUrl,
    current_user: Account = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Corrige l'URL d'un article (ex: lien Google News non résolu).
    Re-scrape le contenu via Trafilatura/Newspaper4k puis relance l'analyse NLP.
    """
    from urllib.parse import urlparse
    from app.services.scraper_service import scraper_service

    article = await db.get(Article, article_id)
    if not article:
        raise HTTPException(status_code=404, detail="Article introuvable")

    new_url = body.url.strip()
    if not new_url.startswith("http"):
        raise HTTPException(status_code=400, detail="URL invalide")

    # Mise à jour de l'URL et du domaine source
    article.url = new_url
    article.source_domain = urlparse(new_url).netloc.lstrip("www.") or article.source_domain

    # Récupérer le keyword pour connaître la langue
    kw = await db.get(Keyword, article.keyword_id)
    language = kw.language if kw else "fr"

    # Re-scrape (Trafilatura → Newspaper4k → Playwright)
    scraped = await scraper_service.scrape(new_url, language=language)
    if scraped.success:
        if scraped.title:
            article.title = scraped.title
        if scraped.content:
            article.content = scraped.content
        if scraped.author:
            article.author = scraped.author
        if scraped.published_at:
            article.published_at = scraped.published_at
        if scraped.image_url:
            article.image_url = scraped.image_url
        if scraped.meta_description:
            article.meta_description = scraped.meta_description
        article.status = ArticleStatus.pending
        article.scraping_error = None

        # Re-analyse NLP uniquement si contenu disponible
        if article.content:
            nlp = await nlp_service.analyze(
                title=article.title or "",
                content=article.content,
                url=new_url,
                keyword=kw.term if kw else "",
                scraper_author=article.author,
                scraper_date=article.published_at,
                meta_description=article.meta_description,
            )
            if not nlp.error:
                article.summary = nlp.summary
                article.summary_en = nlp.summary_en
                article.summary_ar = nlp.summary_ar
                article.tonality = nlp.tonality
                article.tags = nlp.tags
                article.relevance_score = nlp.relevance_score
                article.theme = nlp.theme
                article.weak_signal = nlp.weak_signal
                article.entities_persons = nlp.entities_persons or []
                article.entities_orgs = nlp.entities_orgs or []
                article.entities_places = nlp.entities_places or []
                if nlp.author and not article.author:
                    article.author = nlp.author
                if nlp.published_at and not article.published_at:
                    article.published_at = nlp.published_at

    await db.commit()

    result = await db.execute(
        select(Article).where(Article.id == article.id).options(selectinload(Article.keyword))
    )
    return ArticleOut.from_article(result.scalar_one())


@router.get("/{article_id}", response_model=ArticleOut)
async def get_article(article_id: uuid.UUID, db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(Article).where(Article.id == article_id).options(selectinload(Article.keyword))
    )
    article = result.scalar_one_or_none()
    if not article:
        raise HTTPException(status_code=404, detail="Article introuvable")
    return ArticleOut.from_article(article)


@router.post("/{article_id}/approve")
async def approve_article(
    article_id: uuid.UUID,
    current_user: Account = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Valider un article tel quel."""
    article = await db.get(Article, article_id)
    if not article:
        raise HTTPException(status_code=404, detail="Article introuvable")

    article.status = ArticleStatus.approved
    article.validated_by = current_user.id
    article.validated_at = datetime.utcnow()
    await db.commit()
    return {"status": "approved"}


@router.post("/{article_id}/modify")
async def modify_article(
    article_id: uuid.UUID,
    body: ArticleModify,
    current_user: Account = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Modifier puis valider un article."""
    article = await db.get(Article, article_id)
    if not article:
        raise HTTPException(status_code=404, detail="Article introuvable")

    original = {}
    new_values = {}
    changed_fields = []

    def _to_json(val):
        """Rend une valeur sérialisable JSON pour l'audit log."""
        if isinstance(val, uuid.UUID):
            return str(val)
        if isinstance(val, datetime):
            return val.isoformat()
        return val

    for field, value in body.model_dump(exclude_none=True).items():
        original[field] = _to_json(getattr(article, field, None))
        setattr(article, field, value)
        new_values[field] = _to_json(value)
        changed_fields.append(field)

    article.status = ArticleStatus.modified
    article.validated_by = current_user.id
    article.validated_at = datetime.utcnow()

    # Audit log
    log = ArticleModificationLog(
        article_id=article_id,
        modified_by=current_user.id,
        fields_changed=changed_fields,
        original_values=original,
        new_values=new_values,
    )
    db.add(log)
    await db.commit()
    return {"status": "modified", "fields_changed": changed_fields}


@router.post("/{article_id}/reject")
async def reject_article(
    article_id: uuid.UUID,
    body: ArticleReject,
    current_user: Account = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Rejeter un article."""
    article = await db.get(Article, article_id)
    if not article:
        raise HTTPException(status_code=404, detail="Article introuvable")

    article.status = ArticleStatus.rejected
    article.validated_by = current_user.id
    article.validated_at = datetime.utcnow()
    if body.reason:
        article.reject_reason = body.reason
    await db.commit()
    return {"status": "rejected"}


@router.post("/{article_id}/analyze")
async def analyze_article(
    article_id: uuid.UUID,
    keyword_override: str | None = Query(None, description="Mot-clé pour le calcul de pertinence"),
    db: AsyncSession = Depends(get_db),
):
    """Lance / relance l'analyse NLP sur un article existant (utile après saisie manuelle)."""
    article = await db.get(Article, article_id)
    if not article:
        raise HTTPException(status_code=404, detail="Article introuvable")
    if not article.content:
        raise HTTPException(status_code=400, detail="Aucun contenu à analyser")

    # Récupérer le terme du keyword associé
    from app.models.revue import Keyword as KwModel
    keyword_term = keyword_override or "veille médias Maroc"
    if article.keyword_id:
        kw = await db.get(KwModel, article.keyword_id)
        if kw:
            keyword_term = kw.term

    nlp = await nlp_service.analyze(
        title=article.title or "",
        content=article.content,
        url=article.url,
        keyword=keyword_term,
        scraper_author=article.author,
        scraper_date=article.published_at,
    )

    if nlp.error:
        raise HTTPException(status_code=503, detail=f"NLP error: {nlp.error}")

    # Mise à jour de l'article
    article.summary = nlp.summary
    article.summary_en = nlp.summary_en
    article.summary_ar = nlp.summary_ar
    article.tonality = nlp.tonality
    article.tags = nlp.tags
    article.relevance_score = nlp.relevance_score
    article.theme = nlp.theme
    article.weak_signal = nlp.weak_signal
    article.entities_persons = nlp.entities_persons or []
    article.entities_orgs = nlp.entities_orgs or []
    article.entities_places = nlp.entities_places or []
    if nlp.author and not article.author:
        article.author = nlp.author
    if nlp.published_at and not article.published_at:
        article.published_at = nlp.published_at
    # Titre généré par Claude si l'article n'en avait pas
    if nlp.generated_title and not article.title:
        article.title = nlp.generated_title[:1024]

    await db.commit()
    await db.refresh(article)
    return {
        "status": "analyzed",
        "summary": article.summary,
        "tonality": article.tonality,
        "tags": article.tags,
        "relevance_score": article.relevance_score,
        "entities": {
            "persons": nlp.entities_persons,
            "organizations": nlp.entities_orgs,
            "places": nlp.entities_places,
        },
        "key_themes": nlp.key_themes,
        "market_impact": nlp.market_impact,
        "summaries": {
            "fr": nlp.summary,
            "en": nlp.summary_en,
            "ar": nlp.summary_ar,
        }
    }


@router.post("/{article_id}/manual-entry")
async def manual_entry(
    article_id: uuid.UUID,
    body: ManualEntry,
    current_user: Account = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Saisie manuelle pour un article en erreur de scraping."""
    article = await db.get(Article, article_id)
    if not article:
        raise HTTPException(status_code=404, detail="Article introuvable")
    if article.status != ArticleStatus.error:
        raise HTTPException(status_code=400, detail="L'article n'est pas en erreur")

    article.title = body.title
    article.content = body.content
    article.author = body.author
    article.published_at = body.published_at
    article.summary = body.summary
    article.manually_entered = True
    article.status = ArticleStatus.approved
    article.validated_by = current_user.id
    article.validated_at = datetime.utcnow()

    await db.commit()
    return {"status": "approved", "manually_entered": True}


# ── Retry ──────────────────────────────────────────────────────────────────

@router.post("/{article_id}/retry")
async def retry_article(
    article_id: uuid.UUID,
    current_user: Account = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Re-scrappe et réanalyse un article en erreur."""
    from app.services.scraper_service import scraper_service
    from app.models.revue import Keyword as KwModel

    article = await db.get(Article, article_id)
    if not article:
        raise HTTPException(404, "Article introuvable")
    if article.status not in (ArticleStatus.error, ArticleStatus.pending):
        raise HTTPException(400, "L'article ne peut être réessayé que s'il est en erreur ou en attente")

    # Récupérer langue + terme du keyword
    language = "fr"
    keyword_term = "veille médias"
    if article.keyword_id:
        kw = await db.get(KwModel, article.keyword_id)
        if kw:
            keyword_term = kw.term
            language = kw.language

    # Re-scraping
    scraped = await scraper_service.scrape(article.url, language=language)

    if scraped.error:
        article.scraping_error = scraped.error
        article.status = ArticleStatus.error
        await db.commit()
        return {"status": "error", "error": scraped.error}

    # Mettre à jour le contenu scrappé
    if scraped.title:
        article.title = scraped.title
    if scraped.text:
        article.content = scraped.text
    if scraped.author and not article.author:
        article.author = scraped.author
    if scraped.published_at and not article.published_at:
        article.published_at = scraped.published_at
    article.scraping_error = None

    # Re-NLP si on a du contenu
    if article.content:
        nlp = await nlp_service.analyze(
            title=article.title or "",
            content=article.content,
            url=article.url,
            keyword=keyword_term,
            scraper_author=article.author,
            scraper_date=article.published_at,
        )
        if not nlp.error:
            article.summary = nlp.summary
            article.summary_en = nlp.summary_en
            article.summary_ar = nlp.summary_ar
            article.tonality = nlp.tonality
            article.tags = nlp.tags
            article.relevance_score = nlp.relevance_score
            article.theme = nlp.theme
            article.weak_signal = nlp.weak_signal
            article.entities_persons = nlp.entities_persons or []
            article.entities_orgs = nlp.entities_orgs or []
            article.entities_places = nlp.entities_places or []
            if nlp.author and not article.author:
                article.author = nlp.author
            if nlp.published_at and not article.published_at:
                article.published_at = nlp.published_at
            if nlp.generated_title and not article.title:
                article.title = nlp.generated_title[:1024]

    article.status = ArticleStatus.pending
    await db.commit()
    await db.refresh(article)
    return {"status": "success", "article_id": str(article.id)}


# ── Search ─────────────────────────────────────────────────────────────────

@router.get("/search", response_model=list[ArticleOut])
async def search_articles(
    q: str = Query(..., min_length=2, description="Terme de recherche"),
    revue_id: uuid.UUID | None = Query(None),
    limit: int = Query(20, le=50),
    current_user: Account = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Recherche globale dans les articles (titre, résumé, source, auteur)."""
    from sqlalchemy import or_
    from app.models.client import AccountRole

    term = f"%{q}%"
    search_filter = or_(
        Article.title.ilike(term),
        Article.summary.ilike(term),
        Article.source_domain.ilike(term),
        Article.author.ilike(term),
    )

    stmt = (
        select(Article)
        .where(search_filter)
        .options(selectinload(Article.keyword))
        .order_by(Article.created_at.desc())
        .limit(limit)
    )

    if revue_id:
        stmt = stmt.where(Article.revue_id == revue_id)

    # Restreindre aux revues accessibles selon le rôle
    if current_user.role in (AccountRole.client_admin, AccountRole.client_user):
        from app.models.revue import Revue as RevueModel
        revue_ids_res = await db.execute(
            select(RevueModel.id).where(RevueModel.client_id == current_user.client_id)
        )
        accessible_ids = [r for r, in revue_ids_res]
        stmt = stmt.where(Article.revue_id.in_(accessible_ids))

    result = await db.execute(stmt)
    articles = result.scalars().all()
    return [ArticleOut.from_article(a) for a in articles]


# ── Signalements (flags) ────────────────────────────────────────────────────

class ArticleFlagCreate(BaseModel):
    comment: str | None = None


@router.post("/{article_id}/flag", status_code=201)
async def flag_article(
    article_id: uuid.UUID,
    body: ArticleFlagCreate,
    current_user: Account = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Signaler un article comme non pertinent (côté client)."""
    article = await db.get(Article, article_id)
    if not article:
        raise HTTPException(404, "Article introuvable")

    # Mettre à jour si déjà signalé par cet utilisateur
    existing = await db.scalar(
        select(ArticleFlag).where(
            ArticleFlag.article_id == article_id,
            ArticleFlag.account_id == current_user.id,
        )
    )
    if existing:
        existing.comment = body.comment
        await db.commit()
        return {"success": True, "updated": True}

    flag = ArticleFlag(
        article_id=article_id,
        account_id=current_user.id,
        revue_id=article.revue_id,
        comment=body.comment,
    )
    db.add(flag)
    await db.commit()
    return {"success": True, "created": True}


@router.delete("/{article_id}/flag", status_code=204)
async def unflag_article(
    article_id: uuid.UUID,
    current_user: Account = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Retirer le signalement d'un article."""
    flag = await db.scalar(
        select(ArticleFlag).where(
            ArticleFlag.article_id == article_id,
            ArticleFlag.account_id == current_user.id,
        )
    )
    if flag:
        await db.delete(flag)
        await db.commit()


@router.get("/{article_id}/my-flag")
async def get_my_flag(
    article_id: uuid.UUID,
    current_user: Account = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Vérifie si l'utilisateur courant a signalé cet article."""
    flag = await db.scalar(
        select(ArticleFlag).where(
            ArticleFlag.article_id == article_id,
            ArticleFlag.account_id == current_user.id,
        )
    )
    return {"flagged": flag is not None, "comment": flag.comment if flag else None}


@router.get("/flags/list")
async def list_flags(
    revue_id: uuid.UUID | None = Query(None),
    current_user: Account = Depends(require_admin_plus),
    db: AsyncSession = Depends(get_db),
):
    """Liste tous les signalements — réservé admins."""
    from app.models.client import Account as AccModel

    stmt = (
        select(ArticleFlag, Article, AccModel)
        .join(Article, ArticleFlag.article_id == Article.id)
        .join(AccModel, ArticleFlag.account_id == AccModel.id)
        .order_by(ArticleFlag.created_at.desc())
    )
    if revue_id:
        stmt = stmt.where(ArticleFlag.revue_id == revue_id)

    result = await db.execute(stmt)
    return [
        {
            "id": str(f.id),
            "article_id": str(f.article_id),
            "revue_id": str(f.revue_id),
            "comment": f.comment,
            "created_at": f.created_at.isoformat(),
            "account_name": acc.full_name,
            "account_email": acc.email,
            "article_title": art.title,
            "article_url": art.url,
            "article_source": art.source_domain,
            "article_published_at": art.published_at.isoformat() if art.published_at else None,
        }
        for f, art, acc in result.all()
    ]


@router.get("/admin/rejected", dependencies=[Depends(require_admin_plus)])
async def list_rejected_articles(
    revue_id: uuid.UUID | None = Query(None),
    limit: int = Query(50, le=200),
    db: AsyncSession = Depends(get_db),
):
    """Liste les articles rejetés avec leur motif — réservé admins."""
    from app.models.client import Account as AccModel

    stmt = (
        select(Article, AccModel)
        .outerjoin(AccModel, Article.validated_by == AccModel.id)
        .where(Article.status == ArticleStatus.rejected)
        .order_by(Article.validated_at.desc().nullslast())
        .limit(limit)
    )
    if revue_id:
        stmt = stmt.where(Article.revue_id == revue_id)

    result = await db.execute(stmt)
    rows = result.all()
    return [
        {
            "id": str(art.id),
            "article_id": str(art.id),
            "revue_id": str(art.revue_id),
            "title": art.title,
            "url": art.url,
            "source_domain": art.source_domain,
            "published_at": art.published_at.isoformat() if art.published_at else None,
            "reject_reason": art.reject_reason,
            "validated_at": art.validated_at.isoformat() if art.validated_at else None,
            "validated_by_name": acc.full_name if acc else None,
            "validated_by_email": acc.email if acc else None,
        }
        for art, acc in rows
    ]


@router.get("/revue/{revue_id}/entities-analytics")
async def entities_analytics(
    revue_id: uuid.UUID,
    days: int = Query(30, ge=7, le=365),
    db: AsyncSession = Depends(get_db),
    _=Depends(get_current_user),
):
    """
    Agrégation tags, entités nommées, thèmes pour le reporting.
    Retourne le top 60 tags, top 30 personnes/orgs/lieux, distribution thèmes.
    """
    from datetime import timedelta

    since = datetime.now(timezone.utc) - timedelta(days=days)

    result = await db.execute(
        select(
            Article.tags,
            Article.entities_persons,
            Article.entities_orgs,
            Article.entities_places,
            Article.theme,
            Article.weak_signal,
            Article.key_themes,
            Article.market_impact,
        )
        .where(
            Article.revue_id == revue_id,
            Article.status.in_([ArticleStatus.approved, ArticleStatus.modified]),
            Article.published_at >= since,
        )
        .limit(2000)
    )
    rows = result.all()

    tag_cnt: dict[str, int] = {}
    person_cnt: dict[str, int] = {}
    org_cnt: dict[str, int] = {}
    place_cnt: dict[str, int] = {}
    theme_cnt: dict[str, int] = {}
    key_theme_cnt: dict[str, int] = {}
    weak_signal_total = 0
    market_impacts: list[str] = []

    for r in rows:
        for t in (r.tags or []):
            if t:
                tag_cnt[str(t)] = tag_cnt.get(str(t), 0) + 1
        for p in (r.entities_persons or []):
            name = p.get("name", p) if isinstance(p, dict) else str(p)
            if name:
                person_cnt[name] = person_cnt.get(name, 0) + 1
        for o in (r.entities_orgs or []):
            name = o.get("name", o) if isinstance(o, dict) else str(o)
            if name:
                org_cnt[name] = org_cnt.get(name, 0) + 1
        for pl in (r.entities_places or []):
            name = pl.get("name", pl) if isinstance(pl, dict) else str(pl)
            if name:
                place_cnt[name] = place_cnt.get(name, 0) + 1
        if r.theme:
            theme_cnt[r.theme] = theme_cnt.get(r.theme, 0) + 1
        if r.weak_signal:
            weak_signal_total += 1
        for kt in (r.key_themes or []):
            if kt:
                key_theme_cnt[str(kt)] = key_theme_cnt.get(str(kt), 0) + 1
        if r.market_impact:
            market_impacts.append(r.market_impact)

    def top_n(d: dict, n: int):
        return [{"term": k, "count": v} for k, v in sorted(d.items(), key=lambda x: -x[1])[:n]]

    return {
        "tags":              top_n(tag_cnt, 60),
        "persons":           top_n(person_cnt, 30),
        "orgs":              top_n(org_cnt, 30),
        "places":            top_n(place_cnt, 20),
        "themes":            [{"theme": k, "count": v} for k, v in sorted(theme_cnt.items(), key=lambda x: -x[1])],
        "key_themes":        top_n(key_theme_cnt, 20),
        "market_impacts":    market_impacts[:10],  # 10 derniers impacts marché
        "weak_signal_count": weak_signal_total,
        "total_articles":    len(rows),
    }


@router.get("/revue/{revue_id}/by-entity")
async def articles_by_entity(
    revue_id: uuid.UUID,
    entity_type: str = Query(..., description="tag | person | org | place"),
    entity_value: str = Query(..., description="Valeur exacte de l'entité"),
    limit: int = Query(15, ge=1, le=50),
    db: AsyncSession = Depends(get_db),
    _=Depends(get_current_user),
):
    """Retourne les articles validés mentionnant une entité ou un tag donné."""
    from sqlalchemy.orm import selectinload

    result = await db.execute(
        select(Article)
        .options(selectinload(Article.keyword))
        .where(
            Article.revue_id == revue_id,
            Article.status.in_([ArticleStatus.approved, ArticleStatus.modified]),
        )
        .order_by(Article.published_at.desc())
        .limit(500)
    )
    all_arts = result.scalars().all()

    ev_lower = entity_value.lower()
    matched = []
    for a in all_arts:
        found = False
        if entity_type == "tag":
            found = any(ev_lower == str(t).lower() for t in (a.tags or []))
        elif entity_type == "person":
            for p in (a.entities_persons or []):
                name = p.get("name", p) if isinstance(p, dict) else str(p)
                if ev_lower in name.lower():
                    found = True; break
        elif entity_type == "org":
            for o in (a.entities_orgs or []):
                name = o.get("name", o) if isinstance(o, dict) else str(o)
                if ev_lower in name.lower():
                    found = True; break
        elif entity_type == "place":
            for pl in (a.entities_places or []):
                name = pl.get("name", pl) if isinstance(pl, dict) else str(pl)
                if ev_lower in name.lower():
                    found = True; break
        elif entity_type == "source":
            found = (a.source_domain or "").lower() == ev_lower
        elif entity_type == "author":
            found = (a.author or "").strip().lower() == ev_lower
        if found:
            matched.append(a)
        if len(matched) >= limit:
            break

    return [ArticleOut.from_article(a) for a in matched]


@router.post("/admin/fix-google-news-urls", dependencies=[Depends(require_admin_plus)])
async def fix_google_news_urls(db: AsyncSession = Depends(get_db)):
    """
    Corrige les articles HITL dont l'URL est encore news.google.com (résolution Playwright échouée).
    - Essaie de résoudre chaque URL via Playwright
    - Met à jour url + source_domain si résolu
    - Supprime l'article s'il ne peut pas être résolu (URL inutilisable)
    """
    from app.services.rss_service import _resolve_google_news_url
    from urllib.parse import urlparse

    # Trouver tous les articles HITL avec URL Google News
    result = await db.execute(
        select(Article).where(Article.url.contains("news.google.com"))
    )
    articles = result.scalars().all()

    if not articles:
        return {"message": "Aucun article Google News trouvé", "fixed": 0, "deleted": 0}

    fixed = 0
    deleted = 0

    for art in articles:
        real_url = await _resolve_google_news_url(art.url)
        if "news.google.com" not in real_url and real_url != art.url:
            art.url = real_url
            art.source_domain = urlparse(real_url).netloc.lstrip("www.")
            fixed += 1
        else:
            # Impossible de résoudre → supprimer l'article inutilisable
            await db.delete(art)
            deleted += 1

    await db.commit()
    return {"fixed": fixed, "deleted": deleted, "total": len(articles)}


# ── Alerte critique ───────────────────────────────────────────────────────────

class CriticalAlertBody(BaseModel):
    test_email: str | None = None
    note: str | None = None           # Note optionnelle visible dans le log
    recipients: list[str] | None = None  # Destinataires forcés (surcharge la config)


@router.post("/{article_id}/send-critical")
async def send_critical_alert(
    article_id: str,
    body: CriticalAlertBody = CriticalAlertBody(),
    db: AsyncSession = Depends(get_db),
    user: Account = Depends(require_admin_plus),
):
    """
    Envoie une alerte critique urgente pour un article (email Resend + texte WhatsApp).
    Crée un EmailLog is_critical=True dans la boîte de réception du client.
    """
    from app.core.config import settings
    from app.models.newsletter import EmailLog, NewsletterConfig
    from urllib.parse import urlparse as _urlparse

    aid = uuid.UUID(article_id)
    article = await db.get(Article, aid)
    if not article:
        raise HTTPException(status_code=404, detail="Article introuvable")

    # ── Texte WhatsApp ──────────────────────────────────────────────────
    domain = ""
    try:
        domain = _urlparse(article.url or "").netloc.replace("www.", "")
    except Exception:
        pass
    pub = article.published_at.strftime("%d/%m/%Y") if article.published_at else ""
    meta = " | ".join(filter(None, [domain, pub]))
    summary_txt = (article.summary or "")[:300]
    note_txt = (body.note or "").strip()
    whatsapp_message = (
        f"\U0001f6a8 ALERTE VEILLE MÉDIA\n\n"
        f"{article.title or '(sans titre)'}\n\n"
        + (f"\U0001f4cc {note_txt}\n\n" if note_txt else "")
        + f"{summary_txt}\n\n"
        f"Source : {meta}\n"
        f"Lire : {article.url or ''}"
    )

    if not settings.RESEND_API_KEY:
        return {
            "status": "no_email_service",
            "recipients": [],
            "whatsapp_message": whatsapp_message,
        }

    # ── Email HTML urgent ───────────────────────────────────────────────
    date_str = datetime.now(timezone.utc).strftime("%d/%m/%Y %H:%M")
    html_content = (
        f'<!DOCTYPE html><html lang="fr"><head><meta charset="UTF-8">'
        f'<title>ALERTE — {article.title or ""}</title></head>'
        f'<body style="margin:0;padding:0;background:#f1f5f9;font-family:-apple-system,BlinkMacSystemFont,\'Segoe UI\',sans-serif;">'
        f'<table width="100%" cellpadding="0" cellspacing="0" style="background:#f1f5f9;padding:32px 16px;">'
        f'<tr><td align="center"><table width="600" cellpadding="0" cellspacing="0" '
        f'style="max-width:600px;width:100%;border-radius:12px;overflow:hidden;box-shadow:0 4px 24px rgba(0,0,0,.12);">'
        f'<tr><td style="background:#dc2626;padding:22px 40px;">'
        f'<table width="100%" cellpadding="0" cellspacing="0"><tr>'
        f'<td><p style="margin:0;color:#fff;font-size:11px;font-weight:700;text-transform:uppercase;letter-spacing:2px;">Alerte critique &middot; GMS Veille</p>'
        f'<p style="margin:6px 0 0;color:#fecaca;font-size:11px;">{date_str}</p></td>'
        f'<td align="right"><span style="display:inline-block;background:#fff;color:#dc2626;font-size:11px;font-weight:700;padding:4px 12px;border-radius:4px;letter-spacing:1px;">URGENT</span></td>'
        f'</tr></table></td></tr>'
        f'<tr><td style="background:#fff;padding:32px 40px;">'
        f'<h1 style="margin:0 0 16px;color:#0f172a;font-size:20px;font-weight:700;line-height:1.3;">{article.title or "(sans titre)"}</h1>'
        + (f'<div style="margin:0 0 16px;background:#fef2f2;border-left:3px solid #dc2626;padding:12px 16px;border-radius:4px;"><p style="margin:0;color:#7f1d1d;font-size:13px;font-weight:600;">📌 {note_txt}</p></div>' if note_txt else "")
        + (f'<p style="margin:0 0 16px;color:#475569;font-size:14px;line-height:1.7;text-align:justify;">{summary_txt}</p>' if summary_txt else "")
        + (f'<p style="margin:0 0 20px;color:#94a3b8;font-size:12px;">{meta}</p>' if meta else "")
        + f'<a href="{article.url or "#"}" style="display:inline-block;background:#dc2626;color:#fff;font-size:13px;font-weight:600;text-decoration:none;padding:12px 28px;border-radius:8px;">Lire l\'article complet</a>'
        f'</td></tr>'
        f'<tr><td style="background:#f8fafc;padding:16px 40px;text-align:center;border-top:1px solid #e2e8f0;">'
        f'<p style="margin:0;color:#94a3b8;font-size:11px;">GMS Veille &middot; Alerte critique &middot; {date_str}</p>'
        f'</td></tr></table></td></tr></table></body></html>'
    )

    # ── Destinataires ───────────────────────────────────────────────────
    recipients: list[str] = []
    if body.test_email:
        recipients = [body.test_email]
    elif body.recipients:
        recipients = list(body.recipients)
    elif article.revue_id:
        revue = await db.get(Revue, article.revue_id)
        if revue:
            cfg_res = await db.execute(
                select(NewsletterConfig).where(NewsletterConfig.revue_id == revue.id)
            )
            cfg = cfg_res.scalar_one_or_none()
            if cfg and cfg.include_client_email and revue.client_id:
                client = await db.get(Client, revue.client_id)
                if client and client.email:
                    recipients.append(client.email)
            if cfg:
                for e in (cfg.extra_recipients or []):
                    if e and e not in recipients:
                        recipients.append(e)

    status_send = "sent"
    error_msg = None
    if recipients:
        try:
            import resend
            resend.api_key = settings.RESEND_API_KEY
            resend.Emails.send({
                "from": settings.EMAIL_FROM,
                "to": recipients,
                "subject": f"URGENT \u2014 {article.title or 'Alerte critique'} \u00b7 GMS Veille",
                "html": html_content,
            })
        except Exception as e:
            status_send = "error"
            error_msg = str(e)
    else:
        status_send = "no_recipients"

    # ── Log critique ────────────────────────────────────────────────────
    if article.revue_id:
        log = EmailLog(
            id=uuid.uuid4(),
            revue_id=article.revue_id,
            sent_at=datetime.now(timezone.utc),
            recipients=recipients,
            article_count=1,
            period_from=None,
            period_to=None,
            subject=f"URGENT \u2014 {article.title or 'Alerte critique'}",
            status=status_send if status_send != "no_recipients" else "sent",
            error_message=error_msg,
            triggered_by="critical",
            html_snapshot=html_content,
            article_ids=[str(article.id)],
            is_critical=True,
            read_at=None,
        )
        db.add(log)
        await db.commit()

    return {
        "status": status_send,
        "recipients": recipients,
        "whatsapp_message": whatsapp_message,
        "error": error_msg,
    }
